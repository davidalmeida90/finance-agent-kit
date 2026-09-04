"""Recalculate an Excel workbook so its formulas carry cached values.

Why this exists: openpyxl writes formulas but never their results. A workbook it
produced therefore reads as empty to anything that opens it with
`data_only=True`, which includes most validators and every downstream script.
Opening it once in a real spreadsheet engine and saving fixes that.

Anthropic's dcf-model skill assumes a `recalc.py` that ships with their `xlsx`
skill. That skill's licence forbids redistribution, so this is an independent
implementation with the same job. It also tries Excel before LibreOffice, which
matters on Windows where Excel is usually present and LibreOffice is not.

Usage:
    py -3 recalc.py model.xlsx
    py -3 recalc.py model.xlsx --engine libreoffice
"""
import argparse
import json
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path


def _excel_com(path: Path) -> dict:
    """Windows only. Open in Excel, force a full rebuild, save."""
    try:
        import win32com.client  # type: ignore
    except ImportError:
        return {"ok": False, "engine": "excel", "error": "pywin32 not installed"}

    app = None
    try:
        app = win32com.client.DispatchEx("Excel.Application")
        app.Visible = False
        app.DisplayAlerts = False
        wb = app.Workbooks.Open(str(path.resolve()))
        app.CalculateFullRebuild()
        wb.Save()
        wb.Close(SaveChanges=True)
        return {"ok": True, "engine": "excel"}
    except Exception as e:
        return {"ok": False, "engine": "excel", "error": f"{type(e).__name__}: {e}"}
    finally:
        if app is not None:
            try:
                app.Quit()
            except Exception:
                pass


def _libreoffice(path: Path) -> dict:
    """Cross platform. Convert through LibreOffice, which recalculates on load."""
    soffice = shutil.which("soffice") or shutil.which("soffice.exe")
    if not soffice:
        return {"ok": False, "engine": "libreoffice", "error": "soffice not on PATH"}
    with tempfile.TemporaryDirectory() as td:
        try:
            proc = subprocess.run(
                [soffice, "--headless", "--norestore",
                 "--convert-to", "xlsx:Calc MS Excel 2007 XML", "--outdir", td, str(path.resolve())],
                capture_output=True, text=True, timeout=300,
            )
        except subprocess.TimeoutExpired:
            return {"ok": False, "engine": "libreoffice", "error": "timed out after 300s"}
        produced = Path(td) / path.name
        if not produced.exists():
            return {"ok": False, "engine": "libreoffice",
                    "error": f"no output produced. stderr: {proc.stderr.strip()[:300]}"}
        shutil.copy2(produced, path)
        return {"ok": True, "engine": "libreoffice"}


def _verify(path: Path) -> dict:
    """Count formula cells that now carry a cached value, and any error strings."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {"checked": False, "error": "openpyxl not installed"}

    errs = {"#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "#NULL!", "#NUM!"}
    wf = load_workbook(path, data_only=False)
    wv = load_workbook(path, data_only=True)
    formulas = cached = 0
    found: list[str] = []
    for name in wf.sheetnames:
        sf, sv = wf[name], wv[name]
        for row in sf.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("="):
                    formulas += 1
                    got = sv[c.coordinate].value
                    if got is not None:
                        cached += 1
                    if isinstance(got, str) and got in errs:
                        found.append(f"{name}!{c.coordinate} {got}")
    return {
        "checked": True,
        "formula_cells": formulas,
        "with_cached_values": cached,
        "coverage": f"{cached / formulas:.0%}" if formulas else "n/a",
        "formula_errors": found[:50],
        "error_count": len(found),
    }


def recalc(path: Path, engine: str = "auto") -> dict:
    if not path.exists():
        return {"status": "error", "error": f"file not found: {path}"}

    attempts = []
    order = {"auto": ["excel", "libreoffice"], "excel": ["excel"], "libreoffice": ["libreoffice"]}[engine]
    if sys.platform != "win32" and "excel" in order:
        order = [e for e in order if e != "excel"] or ["libreoffice"]

    result = None
    for eng in order:
        result = _excel_com(path) if eng == "excel" else _libreoffice(path)
        attempts.append(result)
        if result["ok"]:
            break

    out = {"file": str(path), "attempts": attempts}
    if not result or not result["ok"]:
        out["status"] = "error"
        out["hint"] = (
            "Install LibreOffice (winget install TheDocumentFoundation.LibreOffice) "
            "or, on Windows with Excel present, py -3 -m pip install pywin32."
        )
        return out

    out["engine_used"] = result["engine"]
    out["verification"] = _verify(path)
    out["status"] = "error" if out["verification"].get("error_count") else "success"
    return out


def main() -> int:
    ap = argparse.ArgumentParser(description="Recalculate an xlsx so formulas carry cached values.")
    ap.add_argument("file")
    ap.add_argument("--engine", choices=["auto", "excel", "libreoffice"], default="auto")
    a = ap.parse_args()
    res = recalc(Path(a.file), a.engine)
    print(json.dumps(res, indent=2))
    return 0 if res.get("status") == "success" else 1


if __name__ == "__main__":
    raise SystemExit(main())
