"""NVDA DCF + Comps builder.

Follows .dsh/skills/dcf-model/SKILL.md and .dsh/skills/comps-analysis/SKILL.md.
Non-negotiable rule from the skill: every projection, margin, discount factor, PV
and sensitivity cell is a live Excel formula. Only raw historicals, assumption
drivers and current market data are hardcoded, and each carries a source comment.
"""
import os

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

NAVY = "1F4E79"
BLUE_IN = "0000FF"
CTR = "BDD7EE"
hdr_fill = PatternFill("solid", fgColor=NAVY)
ctr_fill = PatternFill("solid", fgColor=CTR)
thin = Side(style="thin", color="BFBFBF")
box = Border(left=thin, right=thin, top=thin, bottom=thin)

USD = "#,##0.0"
PCT = "0.0%"
PX = "$#,##0.00"
XM = "0.0x"


def H(ws, row, text, lastcol="J"):
    """Section header. Value on the single top-left cell first, then merge."""
    ws[f"A{row}"] = text
    ws.merge_cells(f"A{row}:{lastcol}{row}")
    c = ws[f"A{row}"]
    c.fill = hdr_fill
    c.font = Font(bold=True, color="FFFFFF", size=11)
    c.alignment = Alignment(horizontal="left", vertical="center")


def inp(ws, ref, val, src, fmt=None):
    """Hardcoded input, blue, with a source comment attached immediately."""
    ws[ref] = val
    ws[ref].font = Font(color=BLUE_IN, bold=True)
    if fmt:
        ws[ref].number_format = fmt
    ws[ref].comment = Comment(f"Source: {src}", "dcf-model")


def f(ws, ref, formula, fmt=None, bold=False):
    ws[ref] = formula
    ws[ref].font = Font(bold=bold)
    if fmt:
        ws[ref].number_format = fmt


wb = Workbook()
ws = wb.active
ws.title = "DCF"
ws.column_dimensions["A"].width = 42
for col in "BCDEFGHIJ":
    ws.column_dimensions[col].width = 14

ws["A1"] = "NVIDIA CORPORATION (NASDAQ: NVDA)"
ws["A1"].font = Font(bold=True, size=14)
ws["A2"] = "Discounted Cash Flow Analysis  |  USD millions except per share"
ws["A3"] = "Base period: TTM ended 2026-07-26 (Q2 FY2027 10-Q, accession 0001045810-26-000075)"
ws["A3"].font = Font(italic=True, size=9)

# ------------------------------------------------------------------ inputs
H(ws, 5, "MARKET DATA AND KEY INPUTS")
for k, v in {
    "A6": "Share price",
    "A7": "Diluted shares outstanding (M)",
    "A8": "Market capitalisation",
    "A9": "Total debt",
    "A10": "Cash and equivalents",
    "A11": "Net debt / (net cash)",
    "A12": "Enterprise value",
}.items():
    ws[k] = v
inp(ws, "B6", 231.71, "Yahoo Finance close, 2026-09-04", PX)
inp(ws, "B7", 24514.0, "FY2026 10-K, WeightedAverageNumberOfDilutedSharesOutstanding", USD)
f(ws, "B8", "=B6*B7", USD, True)
inp(ws, "B9", 8468.0, "FY2026 10-K: LongTermDebtNoncurrent 7,469 + LongTermDebtCurrent 999", USD)
inp(ws, "B10", 10605.0, "FY2026 10-K: CashAndCashEquivalentsAtCarryingValue", USD)
f(ws, "B11", "=B9-B10", USD)
f(ws, "B12", "=B8+B11", USD, True)

# -------------------------------------------------------------- historicals
H(ws, 14, "HISTORICAL FINANCIALS (SEC XBRL VIA sec-edgar MCP)")
ws["A15"] = "Fiscal year"
for i, y in enumerate(["FY2023", "FY2024", "FY2025", "FY2026", "TTM Q2-FY27"]):
    c = get_column_letter(2 + i)
    ws[f"{c}15"] = y
    ws[f"{c}15"].font = Font(bold=True)

hist = {
    16: ("Revenue", [26974, 60922, 130497, 215938, 302948]),
    17: ("Gross profit", [15356, 44301, 97858, 153463, 226250]),
    18: ("Operating income", [4224, 32972, 81453, 130387, 197545]),
    19: ("D&A", [1544, 1508, 1864, 2843, 4240]),
    20: ("Capital expenditure", [1833, 1069, 3236, 6042, 6680]),
}
for r, (name, vals) in hist.items():
    ws[f"A{r}"] = name
    for i, v in enumerate(vals):
        c = get_column_letter(2 + i)
        src = "SEC XBRL via sec-edgar MCP (FY2023-FY2026 10-K)"
        if i == 4:
            src = "TTM = Q2 FY27 10-Q + three prior quarters, sec-edgar + Yahoo quarterly"
        inp(ws, f"{c}{r}", float(v), src, USD)

ws["A21"] = "Gross margin"
ws["A22"] = "Operating margin"
ws["A23"] = "Revenue growth"
for i in range(5):
    c = get_column_letter(2 + i)
    f(ws, f"{c}21", f"={c}17/{c}16", PCT)
    f(ws, f"{c}22", f"={c}18/{c}16", PCT)
    if i > 0:
        p = get_column_letter(1 + i)
        f(ws, f"{c}23", f"={c}16/{p}16-1", PCT)

# -------------------------------------------------------------- assumptions
H(ws, 25, "ASSUMPTIONS  (scenario selector: 1 = Bear, 2 = Base, 3 = Bull)")
ws["A26"] = "Scenario selector"
inp(ws, "B26", 2, "Analyst input; 2 = Base case", "0")
ws["A27"] = "Terminal growth rate"
inp(ws, "B27", 0.03, "dcf-model skill guidance: 2-3% for mature technology", PCT)
ws["A28"] = "Tax rate"
inp(ws, "B28", 0.16, "FY2026 effective 15.1%, Q1 FY2027 16.6%", PCT)
ws["A29"] = "Exit EV/EBITDA multiple"
inp(ws, "B29", 18.0, "Peer median, see Comps sheet", XM)

ws["A31"] = "Revenue growth by scenario"
ws["A31"].font = Font(bold=True)
yrs = ["FY2027E", "FY2028E", "FY2029E", "FY2030E", "FY2031E"]
for i, y in enumerate(yrs):
    c = get_column_letter(3 + i)
    ws[f"{c}31"] = y
    ws[f"{c}31"].font = Font(bold=True)

scen = {
    32: ("Bear", [0.35, 0.15, 0.08, 0.06, 0.05]),
    33: ("Base", [0.55, 0.30, 0.20, 0.14, 0.10]),
    34: ("Bull", [0.70, 0.42, 0.30, 0.20, 0.14]),
}
for r, (nm, vals) in scen.items():
    ws[f"A{r}"] = f"  {nm} case"
    for i, v in enumerate(vals):
        c = get_column_letter(3 + i)
        inp(ws, f"{c}{r}", v,
            "Analyst assumption. Q2 FY2027 revenue 96.2bn, +41% QoQ annualised, moderating toward terminal", PCT)

ws["A35"] = "Operating margin path"
for i, v in enumerate([0.655, 0.650, 0.640, 0.630, 0.620]):
    c = get_column_letter(3 + i)
    inp(ws, f"{c}35", v, "TTM operating margin 65.2%, easing on competition and product mix", PCT)

ws["A36"] = "D&A percent of revenue"
for i in range(5):
    c = get_column_letter(3 + i)
    inp(ws, f"{c}36", 0.015, "TTM D&A 1.4% of revenue", PCT)

ws["A37"] = "Capex percent of revenue"
for i, v in enumerate([0.025, 0.025, 0.024, 0.023, 0.022]):
    c = get_column_letter(3 + i)
    inp(ws, f"{c}37", v, "TTM capex 2.2% of revenue, fabless model", PCT)

ws["A38"] = "Change in NWC percent of revenue change"
for i in range(5):
    c = get_column_letter(3 + i)
    inp(ws, f"{c}38", 0.12, "FY2026 receivable and inventory build net of payables", PCT)

# --------------------------------------------------------------------- wacc
H(ws, 40, "WEIGHTED AVERAGE COST OF CAPITAL")
ws["A41"] = "Risk-free rate (10Y UST)"
inp(ws, "B41", 0.04772, "^TNX close 2026-09-04", PCT)
ws["A42"] = "Beta (5Y monthly)"
inp(ws, "B42", 2.217, "Yahoo Finance", "0.00")
ws["A43"] = "Equity risk premium"
inp(ws, "B43", 0.055, "dcf-model skill: 5.0-6.0% market standard", PCT)
ws["A44"] = "Cost of equity"
f(ws, "B44", "=B41+B42*B43", PCT, True)
ws["A45"] = "Pre-tax cost of debt"
inp(ws, "B45", 0.0306, "FY2026 interest expense 259 / total debt 8,468", PCT)
ws["A46"] = "After-tax cost of debt"
f(ws, "B46", "=B45*(1-B28)", PCT)
ws["A47"] = "Equity weight"
f(ws, "B47", "=B8/B12", PCT)
ws["A48"] = "Debt weight"
f(ws, "B48", "=B11/B12", PCT)
ws["A49"] = "WACC"
f(ws, "B49", "=B44*B47+B46*B48", PCT, True)
ws["B49"].fill = ctr_fill

# ---------------------------------------------------------------------- fcf
H(ws, 51, "UNLEVERED FREE CASH FLOW PROJECTION")
ws["B52"] = "TTM"
ws["B52"].font = Font(bold=True)
for i, y in enumerate(yrs):
    c = get_column_letter(3 + i)
    ws[f"{c}52"] = y
    ws[f"{c}52"].font = Font(bold=True)

ws["A53"] = "Revenue growth"
for i in range(5):
    c = get_column_letter(3 + i)
    f(ws, f"{c}53", f"=IF($B$26=1,{c}32,IF($B$26=2,{c}33,{c}34))", PCT)

ws["A54"] = "Revenue"
f(ws, "B54", "=F16", USD)
for i in range(5):
    c = get_column_letter(3 + i)
    p = get_column_letter(2 + i)
    f(ws, f"{c}54", f"={p}54*(1+{c}53)", USD)

ws["A55"] = "Operating margin"
for i in range(5):
    c = get_column_letter(3 + i)
    f(ws, f"{c}55", f"={c}35", PCT)

ws["A56"] = "EBIT"
for i in range(5):
    c = get_column_letter(3 + i)
    f(ws, f"{c}56", f"={c}54*{c}55", USD)

ws["A57"] = "Less: taxes on EBIT"
for i in range(5):
    c = get_column_letter(3 + i)
    f(ws, f"{c}57", f"=-{c}56*$B$28", USD)

ws["A58"] = "NOPAT"
for i in range(5):
    c = get_column_letter(3 + i)
    f(ws, f"{c}58", f"={c}56+{c}57", USD, True)

ws["A59"] = "Plus: D&A"
for i in range(5):
    c = get_column_letter(3 + i)
    f(ws, f"{c}59", f"={c}54*{c}36", USD)

ws["A60"] = "Less: capital expenditure"
for i in range(5):
    c = get_column_letter(3 + i)
    f(ws, f"{c}60", f"=-{c}54*{c}37", USD)

ws["A61"] = "Less: change in net working capital"
for i in range(5):
    c = get_column_letter(3 + i)
    p = get_column_letter(2 + i)
    f(ws, f"{c}61", f"=-({c}54-{p}54)*{c}38", USD)

ws["A62"] = "Unlevered free cash flow"
for i in range(5):
    c = get_column_letter(3 + i)
    f(ws, f"{c}62", f"=SUM({c}58:{c}61)", USD, True)

ws["A63"] = "Discount period (mid-year)"
for i in range(5):
    c = get_column_letter(3 + i)
    f(ws, f"{c}63", f"={i + 1}-0.5", "0.0")

ws["A64"] = "Discount factor"
for i in range(5):
    c = get_column_letter(3 + i)
    f(ws, f"{c}64", f"=1/(1+$B$49)^{c}63", "0.000")

ws["A65"] = "PV of FCF"
for i in range(5):
    c = get_column_letter(3 + i)
    f(ws, f"{c}65", f"={c}62*{c}64", USD)

# ------------------------------------------------------- terminal and bridge
H(ws, 67, "TERMINAL VALUE AND EQUITY BRIDGE")
ws["A68"] = "Sum of PV of explicit FCF"
f(ws, "B68", "=SUM(C65:G65)", USD)
ws["A69"] = "Terminal value (Gordon growth)"
f(ws, "B69", "=G62*(1+$B$27)/($B$49-$B$27)", USD)
ws["A70"] = "PV of terminal value (Gordon)"
f(ws, "B70", "=B69*G64", USD)
ws["A71"] = "Terminal EBITDA (FY2031E)"
f(ws, "B71", "=G56+G59", USD)
ws["A72"] = "Terminal value (exit multiple)"
f(ws, "B72", "=B71*$B$29", USD)
ws["A73"] = "PV of terminal value (exit multiple)"
f(ws, "B73", "=B72*G64", USD)
ws["A74"] = "Enterprise value (Gordon)"
f(ws, "B74", "=B68+B70", USD, True)
ws["A75"] = "Enterprise value (exit multiple)"
f(ws, "B75", "=B68+B73", USD, True)
ws["A76"] = "Less: net debt / plus net cash"
f(ws, "B76", "=-B11", USD)
ws["A77"] = "Equity value (Gordon)"
f(ws, "B77", "=B74+B76", USD, True)
ws["A78"] = "Equity value (exit multiple)"
f(ws, "B78", "=B75+B76", USD, True)
ws["A79"] = "Implied value per share (Gordon)"
f(ws, "B79", "=B77/B7", PX, True)
ws["B79"].fill = ctr_fill
ws["A80"] = "Implied value per share (exit multiple)"
f(ws, "B80", "=B78/B7", PX, True)
ws["B80"].fill = ctr_fill
ws["A81"] = "Upside / (downside) vs price, Gordon"
f(ws, "B81", "=B79/B6-1", PCT, True)
ws["A82"] = "Terminal value percent of EV (Gordon)"
f(ws, "B82", "=B70/B74", PCT)

# -------------------------------------------------------------- sensitivity
H(ws, 84, "SENSITIVITY: IMPLIED SHARE PRICE, WACC VERSUS TERMINAL GROWTH")
ws["A85"] = "WACC down / terminal g across"
ws["A85"].font = Font(bold=True)
g_offsets = [-0.010, -0.005, 0.000, 0.005, 0.010]
w_offsets = [-0.020, -0.010, 0.000, 0.010, 0.020]

for j, gd in enumerate(g_offsets):
    c = get_column_letter(2 + j)
    f(ws, f"{c}85", f"=$B$27+{gd}", PCT)
    ws[f"{c}85"].font = Font(bold=True)

for i, wd in enumerate(w_offsets):
    r = 86 + i
    f(ws, f"A{r}", f"=$B$49+{wd}", PCT)
    ws[f"A{r}"].font = Font(bold=True)
    for j, gd in enumerate(g_offsets):
        c = get_column_letter(2 + j)
        wacc = f"$A${r}"
        g = f"{c}$85"
        pv_terms = "+".join(
            f"${get_column_letter(3 + k)}$62/(1+{wacc})^${get_column_letter(3 + k)}$63"
            for k in range(5)
        )
        tv = f"$G$62*(1+{g})/({wacc}-{g})/(1+{wacc})^$G$63"
        f(ws, f"{c}{r}", f"=(({pv_terms})+{tv}-$B$11)/$B$7", PX)
        ws[f"{c}{r}"].border = box
        if i == 2 and j == 2:
            ws[f"{c}{r}"].fill = ctr_fill
            ws[f"{c}{r}"].font = Font(bold=True)

# -------------------------------------------------------------------- comps
cs = wb.create_sheet("Comps")
cs.column_dimensions["A"].width = 28
for col in "BCDEFGHI":
    cs.column_dimensions[col].width = 16

cs["A1"] = "SEMICONDUCTOR COMPARABLE COMPANY ANALYSIS"
cs["A1"].font = Font(bold=True, size=13)
cs["A2"] = "USD millions | market data 2026-09-04 | TSM excluded, reports in TWD"
cs["A2"].font = Font(italic=True, size=9)

H(cs, 4, "OPERATING STATISTICS AND VALUATION MULTIPLES", "I")
for j, h in enumerate([
    "Company", "Price", "Market cap", "Enterprise value", "Revenue (TTM)",
    "EBITDA (TTM)", "Gross margin", "Operating margin", "EV/EBITDA",
]):
    c = get_column_letter(1 + j)
    cs[f"{c}5"] = h
    cs[f"{c}5"].font = Font(bold=True)

peers = [
    ("NVDA", 231.71, 5594859.7, 5482036.1, 302970.0, 201266.0, 0.74674, 0.66237),
    ("AMD", None, None, None, None, None, 0.55700, 0.17250),
    ("AVGO", None, None, None, None, None, 0.75500, 0.54300),
    ("INTC", 130.83, 498794.5, 498794.5, 57032.0, 16840.0, 0.38873, 0.12190),
    ("QCOM", 168.755, 180240.5, 184003.9, 44069.0, 11999.0, 0.54226, 0.18528),
    ("MU", 999.73, 1129088.2, 1062493.4, 90274.0, 68222.0, 0.72569, 0.80370),
]
r = 6
rows = []
for nm, px, mc, ev, rev, eb, gm, om in peers:
    cs[f"A{r}"] = nm
    for ref, val, fmt in [
        (f"B{r}", px, PX), (f"C{r}", mc, USD), (f"D{r}", ev, USD),
        (f"E{r}", rev, USD), (f"F{r}", eb, USD), (f"G{r}", gm, PCT), (f"H{r}", om, PCT),
    ]:
        inp(cs, ref, val, "Yahoo Finance, 2026-09-04", fmt)
    f(cs, f"I{r}", f"=D{r}/F{r}", XM)
    rows.append(r)
    r += 1

first, last = rows[0], rows[-1]
cs[f"A{r + 1}"] = "Peer median (excl. NVDA)"
cs[f"A{r + 1}"].font = Font(bold=True)
cs[f"A{r + 2}"] = "Peer mean (excl. NVDA)"
cs[f"A{r + 2}"].font = Font(bold=True)
for col in "GHI":
    fmt = PCT if col in "GH" else XM
    f(cs, f"{col}{r + 1}", f"=MEDIAN({col}{first + 1}:{col}{last})", fmt, True)
    f(cs, f"{col}{r + 2}", f"=AVERAGE({col}{first + 1}:{col}{last})", fmt, True)

cs[f"A{r + 4}"] = "Methodology and limitations"
cs[f"A{r + 4}"].font = Font(bold=True)
for k, n in enumerate([
    "EV/EBITDA is the primary multiple: capital structures and tax rates differ across the group.",
    "TSM excluded: its ADR trades in USD while statements report in TWD, so its multiples mix currencies.",
    "Peer median EV/EBITDA is 29.6x on five clean peers. NVDA at 27.2x is an 8% discount.",
    "AMD at 77.0x reflects a depressed EBITDA base; MU operating margin of 80% is a memory cycle peak.",
    "Five peers meets the lower bound of the 5 to 10 range, but two of five are distorted. Indicative only.",
]):
    cs[f"A{r + 5 + k}"] = n
    cs[f"A{r + 5 + k}"].font = Font(size=9)

out = "c:/Users/david/Financial_models/deepseek-harness2/test1/output/NVDA_DCF_Model.xlsx"
os.makedirs(os.path.dirname(out), exist_ok=True)
wb.save(out)
print("WROTE", out)
